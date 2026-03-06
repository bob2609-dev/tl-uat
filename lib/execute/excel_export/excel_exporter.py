#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel exporter for TestLink test execution data.
This module creates a Test Suite Execution Summary Excel sheet that matches the UI.
"""

import os
import re
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.label import DataLabelList


class TestExecutionExcelExporter:
    """
    Class for exporting test execution data to Excel.
    """
    def __init__(self, output_dir=None):
        """
        Initialize the Excel exporter.
        
        Args:
            output_dir (str, optional): Directory to save the Excel file
        """
        self.output_dir = output_dir or os.getcwd()
        self.workbook = None
        
        # Define styles
        self.header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Define border style
        self.border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )
        
        # Define status colors
        self.status_colors = {
            'p': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),  # Passed (green)
            'f': PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid'),  # Failed (red)
            'b': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),  # Blocked (yellow)
            'n': PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')   # Not run (light blue)
        }
    
    def create_workbook(self):
        """Create a new Excel workbook."""
        self.workbook = Workbook()
        # Rename the default sheet
        self.workbook.active.title = 'Test Suite Execution Summary'
    
    def apply_header_style(self, cell):
        """Apply header style to a cell."""
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.header_alignment
        cell.border = self.border
        
    def create_test_suite_execution_summary(self, data):
        """
        Create a Test Suite Execution Summary sheet exactly matching the UI format.
        
        Args:
            data (dict): Hierarchical test execution data
        """
        # Special KeyError detection wrapper
        def safe_dict_access(d, key, context="unknown"):
            """Safely access dictionary keys with detailed error logging if KeyError occurs"""
            try:
                return d[key]
            except KeyError as ke:
                error_msg = f"KeyError: '{key}' not found in {context}"
                print(f"\n*** CRITICAL ERROR: {error_msg} ***")
                logging.critical(error_msg)
                # Print more info about the dictionary
                print(f"Dictionary keys: {d.keys() if hasattr(d, 'keys') else 'Not a dict'}")
                print(f"Dictionary type: {type(d)}")
                # Print 2-3 levels of stack trace
                import traceback
                print("\nStack trace:")
                traceback.print_exc()
                # Re-raise the exception with more context
                raise KeyError(f"{key} not found in {context}") from ke
        
        try:
            if not isinstance(data, dict):
                error_msg = f"Error: Expected dict for data in create_test_suite_execution_summary, got {type(data)}"
                print(error_msg)
                logging.error(error_msg)
                self._create_error_sheet(error_msg)
                return
                
            print(f"Creating test suite execution summary with data keys: {data.keys() if data else 'None'}")
            
            # Check metrics_by_suite early
            if 'metrics_by_suite' in data:
                # Inspect each suite for 'count' key
                for suite_id, metrics in data['metrics_by_suite'].items():
                    if metrics and isinstance(metrics, dict):
                        print(f"Suite {suite_id} has keys: {metrics.keys()}")
                        # Directly check for 'count' key
                        if 'count' in metrics:
                            print(f"Suite {suite_id} has 'count': {metrics['count']}")
                        else:
                            print(f"WARNING: Suite {suite_id} is missing 'count' key")
                            # Add a default value to prevent KeyError
                            if 'testcase_count' in metrics:
                                print(f"Using 'testcase_count' instead: {metrics['testcase_count']}")
                                metrics['count'] = metrics['testcase_count']
                            else:
                                print(f"No count key found, setting default count=0")
                                metrics['count'] = 0
                
            sheet = self.workbook.active
            
            # Add title
            sheet['A1'] = 'Test Suite Execution Summary'
            sheet['A1'].font = Font(name='Arial', size=14, bold=True)
            sheet.merge_cells('A1:J1')
            sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Add headers - row 3
            row = 3
            headers = ['Test Paths', 'Test Case Count', 'Passed', 'Failed', 'Blocked', 'Not Run', 
                      'Pass Rate', 'Fail Rate', 'Block Rate', 'Pending Rate']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col)
                cell.value = header
                self.apply_header_style(cell)
            
            # Extract execution data directly matching the UI table
            # This should match what's shown in the screenshot
            try:
                execution_data = self._extract_execution_summary_data(data)
                if not execution_data:  # Empty list
                    print("Warning: No execution data extracted. Adding empty state message to sheet.")
                    sheet.cell(row=row+1, column=1).value = "No execution data available"
                    sheet.merge_cells(f'A{row+1}:J{row+1}')
                    sheet.cell(row=row+1, column=1).alignment = Alignment(horizontal='center')
                    return
            except Exception as e:
                error_msg = f"Error extracting execution summary data: {str(e)}"
                print(error_msg)
                logging.error(error_msg)
                import traceback
                traceback.print_exc()
                sheet.cell(row=row+1, column=1).value = f"Error extracting data: {str(e)}"
                sheet.merge_cells(f'A{row+1}:J{row+1}')
                sheet.cell(row=row+1, column=1).alignment = Alignment(horizontal='center')
                return
            
            # Add data rows
            start_row = row + 1
            for item in execution_data:
                try:
                    # Safely get values with defaults for any missing keys
                    path = item.get('path', 'Unknown Path')
                    total = item.get('total', 0)
                    passed = item.get('passed', 0)
                    failed = item.get('failed', 0)
                    blocked = item.get('blocked', 0)
                    not_run = item.get('not_run', 0)
                    pass_rate = item.get('pass_rate', 0.0)
                    fail_rate = item.get('fail_rate', 0.0)
                    block_rate = item.get('block_rate', 0.0)
                    pending_rate = item.get('pending_rate', 0.0)
                    
                    # Add row data
                    sheet.cell(row=start_row, column=1).value = path
                    sheet.cell(row=start_row, column=2).value = total
                    sheet.cell(row=start_row, column=3).value = passed
                    sheet.cell(row=start_row, column=4).value = failed
                    sheet.cell(row=start_row, column=5).value = blocked
                    sheet.cell(row=start_row, column=6).value = not_run
                    sheet.cell(row=start_row, column=7).value = f"{pass_rate:.2f}%"
                    sheet.cell(row=start_row, column=8).value = f"{fail_rate:.2f}%"
                    sheet.cell(row=start_row, column=9).value = f"{block_rate:.2f}%"
                    sheet.cell(row=start_row, column=10).value = f"{pending_rate:.2f}%"
                    
                    # Apply status colors to values
                    sheet.cell(row=start_row, column=3).fill = self.status_colors['p']
                    sheet.cell(row=start_row, column=4).fill = self.status_colors['f']
                    sheet.cell(row=start_row, column=5).fill = self.status_colors['b']
                    sheet.cell(row=start_row, column=6).fill = self.status_colors['n']
                    
                    # Apply border to all cells
                    for col in range(1, 11):
                        sheet.cell(row=start_row, column=col).border = self.border
                    
                    start_row += 1
                except Exception as e:
                    error_msg = f"Error processing row data: {str(e)}"
                    print(error_msg)
                    logging.error(error_msg)
                    # Continue with next item
                
            # Auto-size columns
            for col in range(1, 11):
                sheet.column_dimensions[get_column_letter(col)].auto_size = True
            sheet.column_dimensions['A'].width = 50  # Test path column wider
                
        except Exception as e:
            error_msg = f"Error in create_test_suite_execution_summary: {str(e)}"
            print(error_msg)
            logging.error(error_msg)
            import traceback
            traceback.print_exc()
            self._create_error_sheet(error_msg)
            return sheet
        
        return sheet

    def _strip_html_tags(self, text):
        """Remove HTML tags from text
        
        Args:
            text (str): Input text with possible HTML tags
            
        Returns:
            str: Text with HTML tags removed
        """
        if not text or not isinstance(text, str):
            return ''

        try:
            # First, remove HTML tags using regex
            clean_text = re.sub(r'<[^>]*>', '', text)
            
            # Replace common HTML entities with their character equivalents
            html_entities = {
                '&lt;': '<',
                '&gt;': '>',
                '&amp;': '&',
                '&quot;': '"',
                '&nbsp;': ' ',
                '&ndash;': '–',
                '&mdash;': '—',
                '&lsquo;': ''',
                '&rsquo;': ''',
                '&ldquo;': '"',
                '&rdquo;': '"',
                '&apos;': "'"
            }
            
            for entity, char in html_entities.items():
                clean_text = clean_text.replace(entity, char)
                
            # Also handle numeric entities (like &#39; for apostrophe)
            clean_text = re.sub(r'&#(\d+);', lambda m: chr(int(m.group(1))), clean_text)
            
            return clean_text.strip()
        except Exception as e:
            # If any error occurs, log it and return the original text with basic tag stripping
            logging.warning(f"HTML stripping error: {str(e)}")
            # Fallback to very basic tag stripping
            return re.sub(r'<[^>]*>', '', text).strip()
    
    def _extract_execution_summary_data(self, data):
        """
        Extract execution summary data for the single sheet.
        
        Args:
            data (dict): Input data structure with metrics from the database
            
        Returns:
            list: List of dictionaries with execution summary data for each path
        """
        try:
            # Add diagnostics to help debug
            print("\nExtracting execution summary data")
            print(f"Data keys: {data.keys() if data else 'None'}")
            
            # Extract data from the metrics_by_suite dictionary
            execution_data = []
            
            # Check if the data structure has metrics_by_suite
            if not data or 'metrics_by_suite' not in data:
                error_msg = "Error: No metrics_by_suite found in data. Cannot generate execution summary."
                print(error_msg)
                logging.error(error_msg)
                return []
            
            print(f"Found metrics_by_suite with {len(data['metrics_by_suite'])} entries")
            
            # Get project and testplan names for path construction
            project_name = data.get('project_name', 'Unknown Project')
            testplan_name = data.get('testplan_name', 'Unknown Plan')
            
            # Process each suite's metrics
            for suite_id, metrics in data['metrics_by_suite'].items():
                try:
                    if not suite_id:
                        print(f"Skipping empty suite_id")
                        continue
                        
                    if not metrics or not isinstance(metrics, dict):
                        print(f"Skipping suite {suite_id} - metrics is {type(metrics)}")
                        continue
                    
                    # Debug each suite's metrics structure
                    print(f"Processing suite {suite_id} with keys: {metrics.keys() if metrics else 'None'}")
                    
                    suite_name = metrics.get('name', 'Unknown Suite')
                    
                    # Try to get count, with fallbacks for all possible keys
                    # This covers both 'count' and 'testcase_count' to handle both formats
                    total = 0
                    for count_key in ['testcase_count', 'count', 'total', 'tc_count']:
                        if count_key in metrics:
                            total = metrics[count_key]
                            print(f"Suite {suite_id}: Found count using key '{count_key}' = {total}")
                            break
                    
                    if total <= 0:
                        print(f"Skipping suite {suite_id} - no test cases (count={total})")
                        continue
                        
                    # Get status counts with safe fallbacks for all possible key formats
                    passed = 0
                    for key in ['passed_count', 'passed', 'p_count', 'p']:
                        if key in metrics:
                            passed = metrics[key]
                            break
                            
                    failed = 0
                    for key in ['failed_count', 'failed', 'f_count', 'f']:
                        if key in metrics:
                            failed = metrics[key]
                            break
                            
                    blocked = 0
                    for key in ['blocked_count', 'blocked', 'b_count', 'b']:
                        if key in metrics:
                            blocked = metrics[key]
                            break
                            
                    not_run = 0
                    for key in ['not_run_count', 'not_run', 'n_count', 'n']:
                        if key in metrics:
                            not_run = metrics[key]
                            break
                    
                    # Calculate rates (percentages) with safeguards
                    executed = passed + failed + blocked
                    pass_rate = (passed / executed * 100) if executed > 0 else 0
                    fail_rate = (failed / executed * 100) if executed > 0 else 0
                    block_rate = (blocked / executed * 100) if executed > 0 else 0
                    pending_rate = (not_run / total * 100) if total > 0 else 0
                    
                    # Use the full path from the database if available, otherwise fall back to simple path
                    path = None
                    for path_key in ['full_path', 'path', 'suite_path']:
                        if path_key in metrics and metrics[path_key]:
                            path = metrics[path_key]
                            print(f"Suite {suite_id}: Found path using key '{path_key}'")
                            break
                    
                    # If no path found, construct it from project/plan/suite names
                    if not path:
                        path = f"{project_name} > {testplan_name} > {suite_name}"
                        print(f"Suite {suite_id}: Using constructed path: {path}")
                    
                    # Add to execution data
                    execution_data.append({
                        'path': path,
                        'total': total,
                        'passed': passed,
                        'failed': failed,
                        'blocked': blocked,
                        'not_run': not_run,
                        'pass_rate': round(pass_rate, 2),
                        'fail_rate': round(fail_rate, 2),
                        'block_rate': round(block_rate, 2),
                        'pending_rate': round(pending_rate, 2)
                    })
                    
                except Exception as suite_error:
                    error_msg = f"Error processing suite {suite_id}: {str(suite_error)}"
                    print(error_msg)
                    logging.error(error_msg)
                    # Continue with next suite
            
            # If no data was found, return an empty list
            if not execution_data:
                print("Warning: No execution data could be extracted from metrics_by_suite")
                return []
            
            # Clean up HTML tags in paths
            for item in execution_data:
                item['path'] = self._strip_html_tags(item['path'])
                
            return execution_data
                
        except Exception as e:
            error_msg = f"Error extracting execution summary data: {str(e)}"
            print(error_msg)
            logging.error(error_msg)
            # Print more detailed debug info
            import traceback
            traceback.print_exc()
            return []
        
        return execution_data
    
    def _extract_test_paths_data(self, data):
        """
        Extract test path data from the hierarchical data.
        
        Args:
            data (dict): Hierarchical test execution data
            
        Returns:
            dict: Dictionary with test paths as keys and status counts as values
        """
        try:
            paths_data = {}
            
            # Debug data structure
            print("\nExtracting test paths data")
            print(f"Data keys: {data.keys() if data else 'None'}")
            
            # Process hierarchical data
            hierarchical_data = data.get('hierarchical_data', {})
            
            if not hierarchical_data:
                print("Warning: No hierarchical_data found in the data structure")
                return paths_data
        
            # Print structure for debugging
            for project_id, project in hierarchical_data.items():
                print(f"Project {project_id} keys: {project.keys() if project else 'None'}")
                project_name = project.get('name', 'Unknown Project')
                
                # Safely access testplans
                testplans = project.get('testplans', {})
                if not isinstance(testplans, dict):
                    print(f"Warning: testplans is not a dictionary for project {project_id}")
                    continue
                    
                for testplan_id, testplan in testplans.items():
                    print(f"Testplan {testplan_id} keys: {testplan.keys() if testplan else 'None'}")
                    testplan_name = testplan.get('name', 'Unknown Test Plan')
                    
                    # Safely access suites
                    suites = testplan.get('suites', {})
                    if not isinstance(suites, dict):
                        print(f"Warning: suites is not a dictionary for testplan {testplan_id}")
                        continue
                        
                    for suite_id, suite in suites.items():
                        print(f"Suite {suite_id} keys: {suite.keys() if suite else 'None'}")
                        suite_name = suite.get('name', 'Unknown Suite')
                        
                        # Strip HTML tags from names
                        project_name = self._strip_html_tags(project_name)
                        testplan_name = self._strip_html_tags(testplan_name)
                        suite_name = self._strip_html_tags(suite_name)
                        
                        # Create the full path as shown in the UI
                        full_path = f"{project_name} > {testplan_name} > {suite_name}"
                        
                        # Initialize counters for this path
                        if full_path not in paths_data:
                            paths_data[full_path] = {
                                'total': 0,
                                'passed': 0,
                                'failed': 0,
                                'blocked': 0,
                                'not_run': 0,
                                'pass_rate': 0.0,
                                'fail_rate': 0.0,
                                'block_rate': 0.0,
                                'pending_rate': 0.0
                            }
                        
                        # Safely access executions
                        executions = suite.get('executions', [])
                        if not isinstance(executions, list):
                            print(f"Warning: executions is not a list for suite {suite_id}")
                            continue
                            
                        # Count executions by status
                        for execution in executions:
                            status = execution.get('execution_status', 'n')  # Default to not run
                            paths_data[full_path]['total'] += 1
                            
                            if status == 'p':
                                paths_data[full_path]['passed'] += 1
                            elif status == 'f':
                                paths_data[full_path]['failed'] += 1
                            elif status == 'b':
                                paths_data[full_path]['blocked'] += 1
                            else:  # status == 'n' or other
                                paths_data[full_path]['not_run'] += 1
                        
                        # Calculate rates
                        total = paths_data[full_path]['total']
                        if total > 0:
                            executed = (paths_data[full_path]['passed'] + 
                                        paths_data[full_path]['failed'] + 
                                        paths_data[full_path]['blocked'])
                            
                            # Calculate rates as shown in UI
                            if executed > 0:
                                paths_data[full_path]['pass_rate'] = (paths_data[full_path]['passed'] / executed) * 100
                                paths_data[full_path]['fail_rate'] = (paths_data[full_path]['failed'] / executed) * 100
                                paths_data[full_path]['block_rate'] = (paths_data[full_path]['blocked'] / executed) * 100
                            
                            # Pending rate is based on not run / total
                            paths_data[full_path]['pending_rate'] = (paths_data[full_path]['not_run'] / total) * 100
                            
            return paths_data
                            
        except Exception as e:
            error_msg = f"Error extracting test paths data: {str(e)}"
            print(error_msg)
            logging.error(error_msg)
            # Print more detailed debug info
            import traceback
            traceback.print_exc()
            return {}
        
        return paths_data
    
    def create_summary_sheet(self, metrics, status_counts):
        """
        Create a summary sheet with overall metrics and charts.
        
        Args:
            metrics (dict): Overall metrics
            status_counts (dict): Status counts
        """
        sheet = self.workbook.create_sheet('Summary')
        
        # Add title
        sheet['A1'] = 'Test Execution Summary'
        sheet['A1'].font = Font(name='Arial', size=14, bold=True)
        sheet.merge_cells('A1:E1')
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add timestamp
        now = datetime.datetime.now()
        sheet['A2'] = f'Generated: {now.strftime("%Y-%m-%d %H:%M:%S")}'
        sheet['A2'].font = Font(name='Arial', size=10, italic=True)
        sheet.merge_cells('A2:E2')
        
        # Add overall metrics table
        row = 4
        sheet.cell(row=row, column=1).value = 'Overall Metrics'
        sheet.cell(row=row, column=1).font = Font(name='Arial', size=12, bold=True)
        sheet.merge_cells(f'A{row}:E{row}')
        
        row += 1
        headers = ['Metric', 'Value', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            self.apply_header_style(cell)
        
        # Add metric rows
        metrics_rows = [
            ('Total Test Cases', metrics.get('testcase_count', 0), ''),
            ('Passed', metrics.get('passed_count', 0), self._format_percentage(metrics.get('passed_count', 0), metrics.get('testcase_count', 0))),
            ('Failed', metrics.get('failed_count', 0), self._format_percentage(metrics.get('failed_count', 0), metrics.get('testcase_count', 0))),
            ('Blocked', metrics.get('blocked_count', 0), self._format_percentage(metrics.get('blocked_count', 0), metrics.get('testcase_count', 0))),
            ('Not Run', metrics.get('not_run_count', 0), self._format_percentage(metrics.get('not_run_count', 0), metrics.get('testcase_count', 0))),
            ('Pass Rate', '', f"{metrics.get('pass_rate', 0):.2f}%")
        ]
        
        start_row = row + 1
        for i, (metric, value, percentage) in enumerate(metrics_rows):
            current_row = start_row + i
            sheet.cell(row=current_row, column=1).value = metric
            sheet.cell(row=current_row, column=2).value = value
            sheet.cell(row=current_row, column=3).value = percentage
            
            # Apply status colors
            if metric == 'Passed':
                sheet.cell(row=current_row, column=1).fill = self.status_colors['p']
            elif metric == 'Failed':
                sheet.cell(row=current_row, column=1).fill = self.status_colors['f']
            elif metric == 'Blocked':
                sheet.cell(row=current_row, column=1).fill = self.status_colors['b']
            elif metric == 'Not Run':
                sheet.cell(row=current_row, column=1).fill = self.status_colors['n']
                
            # Apply border
            for col in range(1, 4):
                sheet.cell(row=current_row, column=col).border = self.border
        
        # Create pie chart for test status
        pie_chart = PieChart()
        pie_chart.title = "Test Execution Status"
        
        # Define data and categories
        status_data = Reference(sheet, min_col=2, min_row=start_row + 1, max_row=start_row + 4)
        status_categories = Reference(sheet, min_col=1, min_row=start_row + 1, max_row=start_row + 4)
        
        pie_chart.add_data(status_data)
        pie_chart.set_categories(status_categories)
        
        # Add data labels
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showPercent = True
        
        # Place chart on sheet
        sheet.add_chart(pie_chart, "G4")
        
        # Auto-size columns
        for col in range(1, 4):
            sheet.column_dimensions[get_column_letter(col)].auto_size = True
        
        return sheet
    
    def create_projects_sheet(self, hierarchical_data):
        """
        Create a sheet with project-level data.
        
        Args:
            hierarchical_data (dict): Hierarchical execution data
        """
        sheet = self.workbook.create_sheet('Projects')
        
        # Add title
        sheet['A1'] = 'Project Summary'
        sheet['A1'].font = Font(name='Arial', size=14, bold=True)
        sheet.merge_cells('A1:G1')
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add headers
        row = 3
        headers = ['Project ID', 'Project Name', 'Test Plans', 'Test Cases', 'Passed', 'Failed', 'Blocked', 'Not Run', 'Pass Rate']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            self.apply_header_style(cell)
        
        # Add project rows
        start_row = row + 1
        for project_id, project_data in hierarchical_data.items():
            project_name = project_data['name']
            testplans = project_data['testplans']
            
            # Calculate project metrics
            testplan_count = len(testplans)
            testcase_count = 0
            status_counts = {'p': 0, 'f': 0, 'b': 0, 'n': 0}
            
            for testplan_id, testplan_data in testplans.items():
                for suite_id, suite_data in testplan_data['suites'].items():
                    for execution in suite_data['executions']:
                        status = execution['execution_status']
                        if status in status_counts:
                            status_counts[status] += 1
                        testcase_count += 1
            
            # Calculate pass rate
            executed_count = status_counts['p'] + status_counts['f'] + status_counts['b']
            pass_rate = (status_counts['p'] / executed_count * 100) if executed_count > 0 else 0
            
            # Add row data
            sheet.cell(row=start_row, column=1).value = project_id
            sheet.cell(row=start_row, column=2).value = project_name
            sheet.cell(row=start_row, column=3).value = testplan_count
            sheet.cell(row=start_row, column=4).value = testcase_count
            sheet.cell(row=start_row, column=5).value = status_counts['p']
            sheet.cell(row=start_row, column=6).value = status_counts['f']
            sheet.cell(row=start_row, column=7).value = status_counts['b']
            sheet.cell(row=start_row, column=8).value = status_counts['n']
            sheet.cell(row=start_row, column=9).value = f"{pass_rate:.2f}%"
            
            # Apply border
            for col in range(1, 10):
                sheet.cell(row=start_row, column=col).border = self.border
            
            start_row += 1
        
        # Auto-size columns
        for col in range(1, 10):
            sheet.column_dimensions[get_column_letter(col)].auto_size = True
        
        return sheet
    
    def create_testplans_sheet(self, hierarchical_data):
        """
        Create a sheet with test plan data.
        
        Args:
            hierarchical_data (dict): Hierarchical execution data
        """
        sheet = self.workbook.create_sheet('TestPlans')
        
        # Add title
        sheet['A1'] = 'Test Plan Summary'
        sheet['A1'].font = Font(name='Arial', size=14, bold=True)
        sheet.merge_cells('A1:H1')
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add headers
        row = 3
        headers = ['Project', 'Test Plan ID', 'Test Plan Name', 'Test Cases', 'Passed', 'Failed', 'Blocked', 'Not Run', 'Pass Rate']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            self.apply_header_style(cell)
        
        # Add test plan rows
        start_row = row + 1
        for project_id, project_data in hierarchical_data.items():
            project_name = project_data['name']
            
            for testplan_id, testplan_data in project_data['testplans'].items():
                testplan_name = testplan_data['name']
                
                # Calculate test plan metrics
                testcase_count = 0
                status_counts = {'p': 0, 'f': 0, 'b': 0, 'n': 0}
                
                for suite_id, suite_data in testplan_data['suites'].items():
                    for execution in suite_data['executions']:
                        status = execution['execution_status']
                        if status in status_counts:
                            status_counts[status] += 1
                        testcase_count += 1
                
                # Calculate pass rate
                executed_count = status_counts['p'] + status_counts['f'] + status_counts['b']
                pass_rate = (status_counts['p'] / executed_count * 100) if executed_count > 0 else 0
                
                # Add row data
                sheet.cell(row=start_row, column=1).value = project_name
                sheet.cell(row=start_row, column=2).value = testplan_id
                sheet.cell(row=start_row, column=3).value = testplan_name
                sheet.cell(row=start_row, column=4).value = testcase_count
                sheet.cell(row=start_row, column=5).value = status_counts['p']
                sheet.cell(row=start_row, column=6).value = status_counts['f']
                sheet.cell(row=start_row, column=7).value = status_counts['b']
                sheet.cell(row=start_row, column=8).value = status_counts['n']
                sheet.cell(row=start_row, column=9).value = f"{pass_rate:.2f}%"
                
                # Apply border
                for col in range(1, 10):
                    sheet.cell(row=start_row, column=col).border = self.border
                
                start_row += 1
        
        # Auto-size columns
        for col in range(1, 10):
            sheet.column_dimensions[get_column_letter(col)].auto_size = True
        
        return sheet
    
    def create_suites_sheet(self, suite_counts):
        """
        Create a sheet with test suite data.
        
        Args:
            suite_counts (dict): Suite counts data
        """
        sheet = self.workbook.create_sheet('TestSuites')
        
        # Add title
        sheet['A1'] = 'Test Suite Summary'
        sheet['A1'].font = Font(name='Arial', size=14, bold=True)
        sheet.merge_cells('A1:G1')
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add headers
        row = 3
        headers = ['Suite ID', 'Suite Name', 'Path', 'Test Cases', 'Passed', 'Failed', 'Blocked', 'Not Run', 'Pass Rate']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            self.apply_header_style(cell)
        
        # Add suite rows
        start_row = row + 1
        for suite_id, suite_data in suite_counts.items():
            # Calculate pass rate
            status_counts = suite_data['statuses']
            executed_count = status_counts['p'] + status_counts['f'] + status_counts['b']
            pass_rate = (status_counts['p'] / executed_count * 100) if executed_count > 0 else 0
            
            # Add row data
            sheet.cell(row=start_row, column=1).value = suite_id
            sheet.cell(row=start_row, column=2).value = suite_data['name']
            sheet.cell(row=start_row, column=3).value = suite_data.get('path', '')
            sheet.cell(row=start_row, column=4).value = suite_data['count']
            sheet.cell(row=start_row, column=5).value = status_counts['p']
            sheet.cell(row=start_row, column=6).value = status_counts['f']
            sheet.cell(row=start_row, column=7).value = status_counts['b']
            sheet.cell(row=start_row, column=8).value = status_counts['n']
            sheet.cell(row=start_row, column=9).value = f"{pass_rate:.2f}%"
            
            # Apply border
            for col in range(1, 10):
                sheet.cell(row=start_row, column=col).border = self.border
            
            start_row += 1
        
        # Auto-size columns
        for col in range(1, 10):
            sheet.column_dimensions[get_column_letter(col)].auto_size = True
        
        return sheet
    
    def create_executions_sheet(self, hierarchical_data):
        """
        Create a sheet with detailed execution data.
        
        Args:
            hierarchical_data (dict): Hierarchical execution data
        """
        sheet = self.workbook.create_sheet('Executions')
        
        # Add title
        sheet['A1'] = 'Test Execution Details'
        sheet['A1'].font = Font(name='Arial', size=14, bold=True)
        sheet.merge_cells('A1:K1')
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add headers
        row = 3
        headers = [
            'Project', 'Test Plan', 'Suite', 'Test Case Name', 'Version', 
            'Build', 'Platform', 'Status', 'Tester', 'Execution Time'
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            self.apply_header_style(cell)
        
        # Add execution rows
        start_row = row + 1
        for project_id, project_data in hierarchical_data.items():
            project_name = project_data['name']
            
            for testplan_id, testplan_data in project_data['testplans'].items():
                testplan_name = testplan_data['name']
                
                for suite_id, suite_data in testplan_data['suites'].items():
                    suite_name = suite_data['name']
                    
                    for execution in suite_data['executions']:
                        # Map status to text
                        status_map = {
                            'p': 'Passed',
                            'f': 'Failed',
                            'b': 'Blocked',
                            'n': 'Not Run'
                        }
                        status = status_map.get(execution['execution_status'], 'Unknown')
                        
                        # Format tester name
                        tester_name = f"{execution['tester_firstname']} {execution['tester_lastname']}" if execution['tester_firstname'] else 'Unknown'
                        
                        # Add row data
                        sheet.cell(row=start_row, column=1).value = project_name
                        sheet.cell(row=start_row, column=2).value = testplan_name
                        sheet.cell(row=start_row, column=3).value = suite_name
                        sheet.cell(row=start_row, column=4).value = execution['tc_name']
                        sheet.cell(row=start_row, column=5).value = execution['tc_version']
                        sheet.cell(row=start_row, column=6).value = execution['build_name']
                        sheet.cell(row=start_row, column=7).value = execution['platform_name'] if execution['platform_name'] else 'None'
                        sheet.cell(row=start_row, column=8).value = status
                        sheet.cell(row=start_row, column=9).value = tester_name
                        sheet.cell(row=start_row, column=10).value = execution['execution_timestamp']
                        
                        # Apply status color to status cell
                        status_cell = sheet.cell(row=start_row, column=8)
                        status_cell.fill = self.status_colors.get(execution['execution_status'], PatternFill())
                        
                        # Apply border
                        for col in range(1, 11):
                            sheet.cell(row=start_row, column=col).border = self.border
                        
                        start_row += 1

    def create_execution_overview_sheet(self, status_counts):
        """
        Create the Execution Overview sheet with summary metrics as shown in the first dashboard section.
        
        Args:
            status_counts (dict): The status counts for all test cases
        """
        sheet = self.workbook.create_sheet('Execution Overview')
        
        # Add title
        sheet['A1'] = 'Test Execution Overview'
        sheet['A1'].font = Font(name='Arial', size=14, bold=True)
        sheet.merge_cells('A1:E1')
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Calculate totals
        passed_count = status_counts.get('p', 0)
        failed_count = status_counts.get('f', 0)
        blocked_count = status_counts.get('b', 0)
        not_run_count = status_counts.get('n', 0)
        total_count = passed_count + failed_count + blocked_count + not_run_count
        executed_count = passed_count + failed_count + blocked_count
        pass_rate = (passed_count / executed_count * 100) if executed_count > 0 else 0
        
        # Create the metrics boxes as in the dashboard
        # Total
        sheet.cell(row=3, column=1).value = 'Total Executions'
        sheet.cell(row=4, column=1).value = total_count
        sheet.cell(row=4, column=1).font = Font(size=20, bold=True)
        
        # Passed
        sheet.cell(row=3, column=2).value = 'Passed'
        sheet.cell(row=4, column=2).value = passed_count
        sheet.cell(row=4, column=2).font = Font(size=20, bold=True, color='00AA00')
        sheet.cell(row=4, column=2).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        
        # Failed
        sheet.cell(row=3, column=3).value = 'Failed'
        sheet.cell(row=4, column=3).value = failed_count
        sheet.cell(row=4, column=3).font = Font(size=20, bold=True, color='AA0000')
        sheet.cell(row=4, column=3).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Blocked
        sheet.cell(row=3, column=4).value = 'Blocked'
        sheet.cell(row=4, column=4).value = blocked_count
        sheet.cell(row=4, column=4).font = Font(size=20, bold=True)
        sheet.cell(row=4, column=4).fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        
        # Not Run
        sheet.cell(row=3, column=5).value = 'Not Run'
        sheet.cell(row=4, column=5).value = not_run_count
        sheet.cell(row=4, column=5).font = Font(size=20, bold=True)
        sheet.cell(row=4, column=5).fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        
        # Pass Rate
        sheet.cell(row=6, column=2).value = 'Pass Rate'
        sheet.cell(row=7, column=2).value = f"{pass_rate:.2f}%"
        sheet.cell(row=7, column=2).font = Font(size=20, bold=True)
        
        # Auto-size columns
        for col in range(1, 6):
            sheet.column_dimensions[get_column_letter(col)].width = 15
        
        return sheet
        
    def create_top_testers_sheet(self, tester_counts):
        """
        Create the Top Testers sheet as shown in the dashboard.
        
        Args:
            tester_counts (dict): Dictionary of tester data keyed by tester name
        """
        sheet = self.workbook.create_sheet('Top Testers')
        
        # Add title
        sheet['A1'] = 'Top Testers'
        sheet['A1'].font = Font(name='Arial', size=14, bold=True)
        sheet.merge_cells('A1:C1')
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add headers
        headers = ['Tester Name', 'Executions']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col)
            cell.value = header
            self.apply_header_style(cell)
        
        # Sort testers by execution count (descending)
        sorted_testers = sorted(
            [(name, data['count']) for name, data in tester_counts.items()],
            key=lambda x: x[1],
            reverse=True
        )
        
        # Add tester data
        row = 4
        for tester_name, count in sorted_testers:
            sheet.cell(row=row, column=1).value = tester_name
            sheet.cell(row=row, column=2).value = count
            row += 1
            
        # Auto-size columns
        for col in range(1, 3):
            sheet.column_dimensions[get_column_letter(col)].width = 20
            
        return sheet
        
    def create_test_suite_progress_sheet(self, suite_counts):
        """
        Create the Test Suite Progress sheet as shown in the dashboard.
        
        Args:
            suite_counts (dict): Dictionary of suite data
        """
        sheet = self.workbook.create_sheet('Test Suite Progress')
        
        # Add title
        sheet['A1'] = 'Test Suite Progress'
        sheet['A1'].font = Font(name='Arial', size=14, bold=True)
        sheet.merge_cells('A1:G1')
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add headers
        row = 3
        headers = ['Test Suite', 'Suite Path', 'Total', 'Passed', 'Failed', 'Blocked', 'Not Run', 'Pass Rate']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            self.apply_header_style(cell)
            
        # Add suite data
        row = 4
        for suite_id, suite_data in suite_counts.items():
            suite_name = suite_data.get('name', 'Unknown Suite')
            suite_path = suite_data.get('path', '')
            total = suite_data.get('count', 0)
            
            # Get status counts
            statuses = suite_data.get('statuses', {})
            passed = statuses.get('p', 0)
            failed = statuses.get('f', 0)
            blocked = statuses.get('b', 0)
            not_run = statuses.get('n', 0)
            
            # Calculate pass rate
            executed = passed + failed + blocked
            pass_rate = (passed / executed * 100) if executed > 0 else 0
            
            # Add row data
            sheet.cell(row=row, column=1).value = suite_name
            sheet.cell(row=row, column=2).value = suite_path
            sheet.cell(row=row, column=3).value = total
            sheet.cell(row=row, column=4).value = passed
            sheet.cell(row=row, column=5).value = failed
            sheet.cell(row=row, column=6).value = blocked
            sheet.cell(row=row, column=7).value = not_run
            sheet.cell(row=row, column=8).value = f"{pass_rate:.2f}%"
            
            # Apply status colors
            sheet.cell(row=row, column=4).fill = self.status_colors['p']
            sheet.cell(row=row, column=5).fill = self.status_colors['f']
            sheet.cell(row=row, column=6).fill = self.status_colors['b']
            sheet.cell(row=row, column=7).fill = self.status_colors['n']
            
            row += 1
            
        # Auto-size columns
        for col in range(1, 9):
            sheet.column_dimensions[get_column_letter(col)].auto_size = True
            
        return sheet
        
    def create_test_suite_execution_summary_sheet(self, hierarchical_data, suite_counts):
        """
        Create the Test Suite Execution Summary sheet as shown in the dashboard.
        
        Args:
            hierarchical_data (dict): Hierarchical test execution data
            suite_counts (dict): Dictionary of suite data
        """
        sheet = self.workbook.create_sheet('Test Suite Execution Summary')
        
        # Add title
        sheet['A1'] = 'Test Suite Execution Summary'
        sheet['A1'].font = Font(name='Arial', size=14, bold=True)
        sheet.merge_cells('A1:J1')
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add headers
        row = 3
        headers = ['Test Path', 'Test Case Count', 'Passed', 'Failed', 'Blocked', 'Not Run', 'Pass Rate', 'Fail Rate', 'Block Rate', 'Pending Rate']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            self.apply_header_style(cell)
        
        # Extract execution summary data
        row = 4
        for project_id, project_data in hierarchical_data.items():
            for testplan_id, testplan_data in project_data['testplans'].items():
                for suite_id, suite_data in testplan_data['suites'].items():
                    # Build full test path
                    test_path = f"{project_data.get('name', 'Unknown Project')} > {testplan_data.get('name', 'Unknown Plan')} > {suite_data.get('name', 'Unknown Suite')}"
                    
                    # Count testcases
                    testcase_count = len(suite_data.get('testcases', []))
                    if testcase_count == 0:
                        continue
                    
                    # Count statuses
                    passed = 0
                    failed = 0
                    blocked = 0
                    not_run = 0
                    
                    for tc in suite_data.get('testcases', []):
                        status = tc.get('execution_status', 'n')
                        if status == 'p':
                            passed += 1
                        elif status == 'f':
                            failed += 1
                        elif status == 'b':
                            blocked += 1
                        else:
                            not_run += 1
                    
                    # Calculate rates
                    total = passed + failed + blocked + not_run
                    executed = passed + failed + blocked
                    pass_rate = (passed / executed * 100) if executed > 0 else 0
                    fail_rate = (failed / executed * 100) if executed > 0 else 0
                    block_rate = (blocked / executed * 100) if executed > 0 else 0
                    pending_rate = (not_run / total * 100) if total > 0 else 0
                    
                    # Add row data
                    sheet.cell(row=row, column=1).value = test_path
                    sheet.cell(row=row, column=2).value = testcase_count
                    sheet.cell(row=row, column=3).value = passed
                    sheet.cell(row=row, column=4).value = failed
                    sheet.cell(row=row, column=5).value = blocked
                    sheet.cell(row=row, column=6).value = not_run
                    sheet.cell(row=row, column=7).value = f"{pass_rate:.2f}%"
                    sheet.cell(row=row, column=8).value = f"{fail_rate:.2f}%"
                    sheet.cell(row=row, column=9).value = f"{block_rate:.2f}%"
                    sheet.cell(row=row, column=10).value = f"{pending_rate:.2f}%"
                    
                    # Apply colors
                    sheet.cell(row=row, column=3).fill = self.status_colors['p']
                    sheet.cell(row=row, column=4).fill = self.status_colors['f']
                    sheet.cell(row=row, column=5).fill = self.status_colors['b']
                    sheet.cell(row=row, column=6).fill = self.status_colors['n']
                    
                    row += 1
        
        # Auto-size columns
        for col in range(1, 11):
            sheet.column_dimensions[get_column_letter(col)].width = 15
        sheet.column_dimensions['A'].width = 50  # Test path column wider
        
        return sheet
        
    def create_execution_details_sheet(self, hierarchical_data):
        """
        Create the Execution Details sheet as shown in the dashboard.
        
        Args:
            hierarchical_data (dict): Hierarchical test execution data
        """
        sheet = self.workbook.create_sheet('Execution Details')
        
        # Add title
        sheet['A1'] = 'Execution Details'
        sheet['A1'].font = Font(name='Arial', size=14, bold=True)
        sheet.merge_cells('A1:F1')
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Setup for hierarchical display
        row = 3
        indent_level = 0
        
        # Process projects
        for project_id, project_data in hierarchical_data.items():
            project_name = project_data.get('name', f"Project {project_id}")
            
            # Add project row
            cell = sheet.cell(row=row, column=1)
            cell.value = project_name
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            sheet.merge_cells(f'A{row}:F{row}')
            row += 1
            
            # Process testplans
            for testplan_id, testplan_data in project_data['testplans'].items():
                testplan_name = testplan_data.get('name', f"Test Plan {testplan_id}")
                
                # Add testplan row with indent
                cell = sheet.cell(row=row, column=1)
                cell.value = f"    {testplan_name}"
                cell.font = Font(bold=True)
                sheet.merge_cells(f'A{row}:F{row}')
                row += 1
                
                # Process suites
                for suite_id, suite_data in testplan_data['suites'].items():
                    suite_name = suite_data.get('name', f"Suite {suite_id}")
                    
                    # Add suite row with indent
                    cell = sheet.cell(row=row, column=1)
                    cell.value = f"        {suite_name}"
                    cell.font = Font(italic=True)
                    sheet.merge_cells(f'A{row}:F{row}')
                    
                    # Add path info to the right (as shown in the screenshot)
                    test_path = suite_data.get('path', '')
                    if test_path:
                        cell = sheet.cell(row=row, column=4)
                        cell.value = test_path
                        cell.font = Font(size=8, color='888888')
                    
                    row += 1
                    
                    # Process testcases (optional, if needed to match dashboard)
                    for tc in suite_data.get('testcases', []):
                        # Skip this if you don't want to show individual testcases
                        pass
        
        # Auto-size columns
        for col in range(1, 7):
            sheet.column_dimensions[get_column_letter(col)].auto_size = True
        sheet.column_dimensions['A'].width = 40  # Name column wider
        
        return sheet
    
    def export_data(self, data, output_file=None):
        """
        Export test execution data to Excel with two sheets:
        1. Test Suite Execution Summary sheet matching the UI
        2. Test Suite Progress sheet showing progress metrics by suite
        
        Args:
            data (dict): The test execution data, including suite_progress
            output_file (str, optional): Output file path
            
        Returns:
            str: Path to the created Excel file
        """
        try:
            # Validate input data
            if not isinstance(data, dict):
                error_msg = f"Error: Expected dict for data, got {type(data)}"
                print(error_msg)
                logging.error(error_msg)
                return None
            
            # Debug data structure
            print(f"Export data keys: {data.keys() if data else 'None'}")
            
            if not output_file:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_file = f"testlink_export_{timestamp}.xlsx"
                
            if not os.path.isabs(output_file):
                output_file = os.path.join(self.output_dir, output_file)
            
            # Create workbook and sheets
            self.create_workbook()
            
            try:
                # Create the main Test Suite Execution Summary sheet
                self.create_test_suite_execution_summary(data)
                
                # Check if we have suite progress data for the second sheet
                if 'suite_progress' in data and data['suite_progress']:
                    logging.info(f"Creating Test Suite Progress sheet with {len(data['suite_progress'])} suites")
                    self.create_test_suite_progress_sheet(data['suite_progress'])
                    logging.info("Test Suite Progress sheet created successfully")
                else:
                    logging.warning("No suite_progress data found for second sheet")
                    
            except Exception as e:
                error_msg = f"Error creating Excel sheets: {str(e)}"
                print(error_msg)
                logging.error(error_msg)
                import traceback
                traceback.print_exc()
                # Create a minimal sheet with error message
                self._create_error_sheet(f"Failed to create Excel: {str(e)}")
            
            # Save workbook
            self.workbook.save(output_file)
            logging.info(f"Excel file saved successfully to {output_file}")
            return output_file
        
        except Exception as e:
            error_msg = f"Error exporting data: {str(e)}"
            print(error_msg)
            logging.error(error_msg)
            import traceback
            traceback.print_exc()
            return None
    
    def create_suite_execution_summary_sheet(self, hierarchical_data, suite_counts):
        """
        Create a Test Suite Execution Summary sheet that matches the UI format.
        
        Args:
            hierarchical_data (dict): Hierarchical execution data
            suite_counts (dict): Suite counts data
        """
        sheet = self.workbook.create_sheet('Test Suite Execution Summary')
    
        # Add title
        sheet['A1'] = 'Test Suite Execution Summary'
        sheet['A1'].font = Font(name='Arial', size=14, bold=True)
        sheet.merge_cells('A1:K1')
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
        # Add headers
        row = 3
        headers = ['Test Paths', 'Test Case Count', 'Passed', 'Failed', 'Blocked', 'Not Run', 'Pass Rate', 'Fail Rate', 'Block Rate', 'Pending Rate']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            self.apply_header_style(cell)
            
        # Create a dictionary to hold full path data from hierarchical data
        path_data = {}
        
        # Extract test paths from hierarchical data
        for project_id, project_data in hierarchical_data.items():
            project_name = project_data['name']
            
            for testplan_id, testplan_data in project_data['testplans'].items():
                testplan_name = testplan_data['name']
                
                for suite_id, suite_data in testplan_data['suites'].items():
                    suite_name = suite_data['name']
                    
                    # Build the full path
                    full_path = f"{project_name} > {testplan_name} > {suite_name}"
                    
                    # Count test cases and statuses for this path
                    if full_path not in path_data:
                        path_data[full_path] = {
                            'count': 0,
                            'statuses': {'p': 0, 'f': 0, 'b': 0, 'n': 0}
                        }
                        
                    for execution in suite_data['executions']:
                        status = execution['execution_status']
                        if status in path_data[full_path]['statuses']:
                            path_data[full_path]['statuses'][status] += 1
                            path_data[full_path]['count'] += 1
        
        # Add data rows
        start_row = row + 1
        for path, data in path_data.items():
            count = data['count']
            statuses = data['statuses']
            
            # Calculate rates
            executed = statuses['p'] + statuses['f'] + statuses['b']
            pass_rate = (statuses['p'] / executed * 100) if executed > 0 else 0
            fail_rate = (statuses['f'] / executed * 100) if executed > 0 else 0
            block_rate = (statuses['b'] / executed * 100) if executed > 0 else 0
            pending_rate = (statuses['n'] / count * 100) if count > 0 else 100  # Not run % is pending rate
            
            # Add row data
            sheet.cell(row=start_row, column=1).value = path
            sheet.cell(row=start_row, column=2).value = count
            sheet.cell(row=start_row, column=3).value = statuses['p']
            sheet.cell(row=start_row, column=4).value = statuses['f']
            sheet.cell(row=start_row, column=5).value = statuses['b']
            sheet.cell(row=start_row, column=6).value = statuses['n']
            sheet.cell(row=start_row, column=7).value = f"{pass_rate:.2f}%"
            sheet.cell(row=start_row, column=8).value = f"{fail_rate:.2f}%"
            sheet.cell(row=start_row, column=9).value = f"{block_rate:.2f}%"
            sheet.cell(row=start_row, column=10).value = f"{pending_rate:.2f}%"
            
            # Apply status colors to values
            sheet.cell(row=start_row, column=3).fill = self.status_colors['p']
            sheet.cell(row=start_row, column=4).fill = self.status_colors['f']
            sheet.cell(row=start_row, column=5).fill = self.status_colors['b']
            sheet.cell(row=start_row, column=6).fill = self.status_colors['n']
            
            # Apply border to all cells
            for col in range(1, 11):
                sheet.cell(row=start_row, column=col).border = self.border
            
            start_row += 1
            
        # Auto-size columns
        for col in range(1, 11):
            sheet.column_dimensions[get_column_letter(col)].auto_size = True
            
        return sheet

def export_data(self, data, output_file=None):
    """
    Export test execution data to an Excel file that matches the TestLink dashboard sections.
    
    Args:
        data (dict): The test execution data
        output_file (str, optional): Output file path. If provided, this is used as is.
                                     If not provided, a file is created in self.output_dir.
        
    Returns:
        str: Path to the created Excel file
    """
    # If output_file is not provided, generate a name in the output directory
    if not output_file:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"test_execution_summary_{timestamp}.xlsx"
        output_file = os.path.join(self.output_dir, filename)
        
    # Ensure the directory exists
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        logging.info(f"Created directory: {output_dir}")
        
    self.workbook = Workbook()
    
    # Get data components
    hierarchical_data = data.get('hierarchical_data', {})
    status_counts = data.get('status_counts', {})
    tester_counts = data.get('tester_counts', {})
    suite_counts = data.get('suite_counts', {})
    
    # Create sheets that match the TestLink dashboard sections
    self.create_execution_overview_sheet(status_counts)
    self.create_top_testers_sheet(tester_counts)
    self.create_test_suite_progress_sheet(suite_counts)
    self.create_test_suite_execution_summary_sheet(hierarchical_data, suite_counts)
    self.create_suite_execution_summary_sheet(hierarchical_data, suite_counts)
    self.create_execution_details_sheet(hierarchical_data)
    
    # Remove default sheet
    if 'Sheet' in self.workbook.sheetnames:
        self.workbook.remove(self.workbook['Sheet'])
    
    # Save workbook
    self.workbook.save(output_file)
    return output_file

def _format_percentage(self, numerator, denominator):
    """Format a percentage."""
    if denominator == 0:
        return '0.00%'
    return f"{(numerator / denominator * 100):.2f}%"

    def _create_error_sheet(self, error_message):
        """
        Create a simple error sheet when data extraction fails.
        
        Args:
            error_message (str): The error message to display
        """
        # Ensure workbook exists
        if not self.workbook:
            self.create_workbook()
            
        # Get active sheet
        sheet = self.workbook.active
        sheet.title = "Error"
        
        # Add error message
        sheet['A1'] = "Error Generating Test Execution Summary"
        sheet['A1'].font = Font(name='Arial', size=14, bold=True, color='FF0000')
        sheet.merge_cells('A1:D1')
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        sheet['A3'] = "An error occurred while processing the test execution data:"
        sheet['A3'].font = Font(name='Arial', size=12, bold=True)
        
        sheet['A4'] = error_message
        sheet['A4'].font = Font(name='Arial', size=11)
        
        # Add timestamp
        sheet['A6'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        sheet['A6'].font = Font(name='Arial', size=10, italic=True)
        
        # Auto-size columns
        for col in range(1, 5):
            sheet.column_dimensions[get_column_letter(col)].auto_size = True
